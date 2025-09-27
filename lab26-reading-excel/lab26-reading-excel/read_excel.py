import openpyxl

# Load workbook
workbook = openpyxl.load_workbook("mydata.xlsx")
sheet = workbook.active

# Read a single cell
cell_value = sheet['A1'].value
print(f"The value in cell A1 is: {cell_value}")

# Read all rows
print("\nReading all rows:")
for row in sheet.iter_rows(min_row=1, max_col=2, max_row=sheet.max_row, values_only=True):
    print(row)

# Calculate sum of ages in column B
sum_age = 0
for row in sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row, values_only=True):
    sum_age += row[1]

print(f"\nThe total sum of ages is: {sum_age}")
