from openpyxl import Workbook

# Create workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# Add headers
ws['A1'] = "Name"
ws['B1'] = "Age"

# Add rows
ws.append(["Alice", 25])
ws.append(["Bob", 30])

# Save file
wb.save("mydata.xlsx")

print("mydata.xlsx created successfully!")
